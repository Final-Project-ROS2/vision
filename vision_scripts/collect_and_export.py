#!/usr/bin/env python3
"""
Vision Pipeline Data Collector & Excel Exporter

Calls all 4 vision services in sequence, stores the last 20 runs in a JSON
history file, and exports every run to an Excel workbook (4 sheets).

Services called (in order):
    1. /vision/run_pipeline          (std_srvs/Trigger)  -> SAM detects + publishes
    2. /vision/classify_bbox_filtered (std_srvs/Trigger)  -> CLIP labels per region
    3. /vision/understand_scene      (std_srvs/Trigger)   -> objects + relations + grasps

Output files (written to workspace root next to README.md):
    vision_runs_history.json   — rolling last-20-runs database
    vision_runs_export.xlsx    — Excel workbook, 4 sheets:
        • Runs      — one row per pipeline run
        • Objects   — one row per detected object
        • Relations — one row per spatial relation
        • Grasps    — one row per grasp pose

Usage:
    # Make sure the vision nodes are already running, then:
    python3 vision_scripts/collect_and_export.py

    # Or as a ROS2 entry point (after colcon build):
    ros2 run vision collect_and_export
"""

import math
import rclpy
from rclpy.node import Node
from std_srvs.srv import Trigger
import json
import time
import os
import sys
from datetime import datetime
from pathlib import Path

# OBB service uses custom_interfaces (optional — graceful skip if not built)
try:
    from custom_interfaces.srv import FindObjectAngle
    _OBB_AVAILABLE = True
except ImportError:
    _OBB_AVAILABLE = False

# ---------------------------------------------------------------------------
# Output file paths  (workspace root = parent of this script's package dir)
# ---------------------------------------------------------------------------
_SCRIPT_DIR = Path(__file__).resolve().parent
_WORKSPACE_ROOT = _SCRIPT_DIR.parent
HISTORY_FILE = _WORKSPACE_ROOT / "vision_runs_history.json"
EXCEL_FILE = _WORKSPACE_ROOT / "vision_runs_export.xlsx"
MAX_HISTORY = 20


# ---------------------------------------------------------------------------
# Helper: load / save history
# ---------------------------------------------------------------------------

def _load_history():
    if HISTORY_FILE.exists():
        try:
            with open(HISTORY_FILE, "r") as f:
                data = json.load(f)
            if isinstance(data, list):
                return data
        except Exception:
            pass
    return []


def _save_history(runs):
    runs = runs[-MAX_HISTORY:]  # keep last 20 only
    with open(HISTORY_FILE, "w") as f:
        json.dump(runs, f, indent=2)


# ---------------------------------------------------------------------------
# Helper: export to Excel using openpyxl
# ---------------------------------------------------------------------------

def _export_excel(runs):
    """Write runs list → Excel with 4 sheets."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("[WARN] openpyxl not installed. Run: pip install openpyxl")
        print("[WARN] Skipping Excel export. JSON history saved successfully.")
        return

    wb = Workbook()

    # ---- colour scheme ----
    header_fill = PatternFill("solid", fgColor="4F81BD")
    header_font = Font(bold=True, color="FFFFFF")
    alt_fill    = PatternFill("solid", fgColor="DCE6F1")

    def _write_sheet(ws, headers, rows, colour_col=None):
        """Write header row then data rows with alternating row colour."""
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for row_idx, row in enumerate(rows, start=2):
            ws.append(row)
            if row_idx % 2 == 0:
                for cell in ws[row_idx]:
                    cell.fill = alt_fill
        # Auto-width
        for col_idx, header in enumerate(headers, start=1):
            col_letter = get_column_letter(col_idx)
            max_len = len(str(header))
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    try:
                        max_len = max(max_len, len(str(cell.value or "")))
                    except Exception:
                        pass
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    # ---- Sheet 1: Runs ----
    ws_runs = wb.active
    ws_runs.title = "Runs"
    runs_headers = [
        "Run No", "Timestamp", "SAM Total Objects", "CLIP Filtered Regions",
        "Graspable Objects", "Total Relations", "Avg SAM Confidence",
        "Avg IoU", "Stability Rate (%)", "Scene Description",
        "SAM Success", "CLIP Success", "Scene Success", "OBB Success",
        "Total Latency (s)", "SAM Latency (s)", "CLIP Latency (s)",
        "Scene Latency (s)", "OBB Latency (s)"
    ]
    runs_rows = []
    for run in runs:
        meta = run.get("meta", {})
        sam  = run.get("sam", {})
        clip = run.get("clip", {})
        scene = run.get("scene", {})
        obb   = run.get("obb", {})
        runs_rows.append([
            meta.get("run_no", ""),
            meta.get("timestamp", ""),
            sam.get("total_detections", ""),
            clip.get("filtered_regions", ""),
            scene.get("graspable_objects", ""),
            scene.get("total_relations", ""),
            sam.get("avg_confidence", ""),
            sam.get("average_iou", ""),
            round(sam.get("stability_rate", 0) * 100, 1),
            scene.get("scene_description", ""),
            sam.get("success", ""),
            clip.get("success", ""),
            scene.get("success", ""),
            obb.get("success", ""),
            meta.get("latency_s", ""),
            sam.get("latency_s", ""),
            clip.get("latency_s", ""),
            scene.get("latency_s", ""),
            obb.get("latency_s", ""),
        ])
    _write_sheet(ws_runs, runs_headers, runs_rows)

    # ---- Sheet 2: Objects ----
    ws_obj = wb.create_sheet("Objects")
    obj_headers = [
        "Run No", "Timestamp", "Object ID", "CLIP Label",
        "BBox X1", "BBox Y1", "BBox X2", "BBox Y2",
        "SAM Confidence", "CLIP Confidence",
        "Distance (cm)", "IoU Score", "Is Stable",
        "Has Grasp", "Grasp Quality", "Grasp Width (m)",
        # OBB columns
        "OBB Angle (deg)", "OBB Theta (rad)",
        "OBB Width (px)", "OBB Height (px)",
        "OBB Center U", "OBB Center V",
    ]
    obj_rows = []
    for run in runs:
        meta  = run.get("meta", {})
        run_no = meta.get("run_no", "")
        ts     = meta.get("timestamp", "")
        for obj in run.get("objects", []):
            grasp = obj.get("grasp", {})
            obj_rows.append([
                run_no, ts,
                obj.get("object_id", ""),
                obj.get("label", ""),
                obj.get("bbox_x1", ""), obj.get("bbox_y1", ""),
                obj.get("bbox_x2", ""), obj.get("bbox_y2", ""),
                obj.get("sam_confidence", ""),
                obj.get("clip_confidence", ""),
                obj.get("distance_cm", ""),
                obj.get("iou_score", ""),
                obj.get("is_stable", ""),
                obj.get("has_grasp", ""),
                grasp.get("quality_score", ""),
                grasp.get("width_m", ""),
                # OBB
                obj.get("obb_angle_deg", ""),
                obj.get("obb_theta_rad", ""),
                obj.get("obb_width_px",  ""),
                obj.get("obb_height_px", ""),
                obj.get("obb_center_u",  ""),
                obj.get("obb_center_v",  ""),
            ])
    _write_sheet(ws_obj, obj_headers, obj_rows)

    # ---- Sheet 3: Relations ----
    ws_rel = wb.create_sheet("Relations")
    rel_headers = [
        "Run No", "Timestamp", "Scene ID",
        "Subject", "Relation", "Target Object",
        "Confidence", "Distance 2D", "Description"
    ]
    rel_rows = []
    for run in runs:
        meta   = run.get("meta", {})
        run_no = meta.get("run_no", "")
        ts     = meta.get("timestamp", "")
        scene  = run.get("scene", {})
        scene_id = scene.get("scene_id", "")
        for rel in run.get("relations", []):
            rel_rows.append([
                run_no, ts, scene_id,
                rel.get("subject", ""),
                rel.get("relation", ""),
                rel.get("target_object", ""),
                rel.get("confidence", ""),
                rel.get("distance_2d", ""),
                rel.get("description", ""),
            ])
    _write_sheet(ws_rel, rel_headers, rel_rows)

    # ---- Sheet 4: Grasps ----
    ws_grasp = wb.create_sheet("Grasps")
    grasp_headers = [
        "Run No", "Timestamp",
        "Object ID", "Pos X (m)", "Pos Y (m)", "Pos Z (m)",
        "Orient X", "Orient Y", "Orient Z", "Orient W",
        "Quality Score", "Grasp Width (m)", "Approach Direction"
    ]
    grasp_rows = []
    for run in runs:
        meta   = run.get("meta", {})
        run_no = meta.get("run_no", "")
        ts     = meta.get("timestamp", "")
        for g in run.get("grasps", []):
            grasp_rows.append([
                run_no, ts,
                g.get("object_id", ""),
                g.get("pos_x", ""), g.get("pos_y", ""), g.get("pos_z", ""),
                g.get("orient_x", ""), g.get("orient_y", ""),
                g.get("orient_z", ""), g.get("orient_w", ""),
                g.get("quality_score", ""),
                g.get("width_m", ""),
                g.get("approach_direction", ""),
            ])
    _write_sheet(ws_grasp, grasp_headers, grasp_rows)

    # ---- Sheet 5: OBB Angles ----
    ws_obb = wb.create_sheet("OBB Angles")
    obb_headers = [
        "Run No", "Timestamp",
        "Object ID", "CLIP Label",
        "OBB Angle (deg)", "OBB Theta (rad)",
        "OBB Width (px)", "OBB Height (px)",
        "OBB Center U", "OBB Center V",
        "BBox X1", "BBox Y1", "BBox X2", "BBox Y2",
        "SAM Confidence", "Distance (cm)",
        "OBB Latency (s)",
    ]
    obb_rows = []
    for run in runs:
        meta   = run.get("meta", {})
        run_no = meta.get("run_no", "")
        ts     = meta.get("timestamp", "")
        obb_lat = run.get("obb", {}).get("latency_s", "")
        for obj in run.get("objects", []):
            if obj.get("obb_angle_deg", "") == "":
                continue  # skip objects where OBB wasn't available
            obb_rows.append([
                run_no, ts,
                obj.get("object_id", ""),
                obj.get("label", ""),
                obj.get("obb_angle_deg", ""),
                obj.get("obb_theta_rad", ""),
                obj.get("obb_width_px",  ""),
                obj.get("obb_height_px", ""),
                obj.get("obb_center_u",  ""),
                obj.get("obb_center_v",  ""),
                obj.get("bbox_x1", ""), obj.get("bbox_y1", ""),
                obj.get("bbox_x2", ""), obj.get("bbox_y2", ""),
                obj.get("sam_confidence", ""),
                obj.get("distance_cm", ""),
                obb_lat,
            ])
    _write_sheet(ws_obb, obb_headers, obb_rows)

    wb.save(str(EXCEL_FILE))
    print(f"[OK] Excel exported → {EXCEL_FILE}")


# ---------------------------------------------------------------------------
# ROS2 collector node
# ---------------------------------------------------------------------------

class VisionDataCollector(Node):

    def __init__(self):
        super().__init__("vision_data_collector")

        self._sam_client   = self.create_client(Trigger, "/vision/run_pipeline")
        self._clip_client  = self.create_client(Trigger, "/vision/classify_bbox_filtered")
        self._scene_client = self.create_client(Trigger, "/vision/understand_scene")

        # OBB client — only if custom_interfaces is built
        if _OBB_AVAILABLE:
            self._obb_client = self.create_client(FindObjectAngle, "/obb/find_object_angle")
        else:
            self._obb_client = None
            self.get_logger().warn("custom_interfaces not found — OBB step will be skipped")

    # ------------------------------------------------------------------
    # Low-level call helper
    # ------------------------------------------------------------------

    def _call(self, client, service_name, timeout=10.0):
        """Wait for service, call it, return (success, message_str)."""
        self.get_logger().info(f"Waiting for {service_name} ...")
        if not client.wait_for_service(timeout_sec=5.0):
            self.get_logger().warn(f"{service_name} not available — skipping")
            return False, None

        future = client.call_async(Trigger.Request())
        start = time.time()
        while not future.done():
            rclpy.spin_once(self, timeout_sec=0.05)
            if time.time() - start > timeout:
                self.get_logger().error(f"{service_name} call timed out")
                return False, None

        result = future.result()
        if result is None:
            return False, None
        return result.success, result.message

    def _call_obb(self, timeout=10.0):
        """Call /obb/find_object_angle (FindObjectAngle service). Returns response or None."""
        if self._obb_client is None:
            return None
        service_name = "/obb/find_object_angle"
        self.get_logger().info(f"Waiting for {service_name} ...")
        if not self._obb_client.wait_for_service(timeout_sec=5.0):
            self.get_logger().warn(f"{service_name} not available — skipping OBB step")
            return None
        future = self._obb_client.call_async(FindObjectAngle.Request())
        start = time.time()
        while not future.done():
            rclpy.spin_once(self, timeout_sec=0.05)
            if time.time() - start > timeout:
                self.get_logger().error(f"{service_name} call timed out")
                return None
        return future.result()

    # ------------------------------------------------------------------
    # Collect one run
    # ------------------------------------------------------------------

    def collect_run(self, run_no):
        """Call all services and return a unified run dict."""
        ts = datetime.utcnow().isoformat() + "Z"
        t0 = time.perf_counter()

        run = {
            "meta": {"run_no": run_no, "timestamp": ts},
            "sam":  {"success": False, "latency_s": 0.0},
            "clip": {"success": False, "latency_s": 0.0},
            "scene": {"success": False, "latency_s": 0.0},
            "obb":  {"success": False, "total_objects": 0, "latency_s": 0.0},
            "objects":   [],
            "relations": [],
            "grasps":    [],
        }

        # ---- 1. SAM ----
        self.get_logger().info("=" * 60)
        self.get_logger().info(f"RUN #{run_no} — Step 1: SAM /vision/run_pipeline")
        t_sam = time.perf_counter()
        sam_ok, sam_msg = self._call(self._sam_client, "/vision/run_pipeline")
        run["sam"]["latency_s"] = round(time.perf_counter() - t_sam, 3)
        if sam_ok and sam_msg:
            try:
                sam_data = json.loads(sam_msg)
                summary  = sam_data.get("summary", {})
                metrics  = sam_data.get("metrics", {})
                coco     = metrics.get("coco_ap_style", {})
                circ     = metrics.get("circularity_confidence", {})

                run["sam"] = {
                    "success":       True,
                    "latency_s":     run["sam"]["latency_s"],
                    "total_detections": summary.get("total_detections", 0),
                    "avg_confidence":   circ.get("average_confidence", 0.0),
                    "average_iou":      coco.get("average_iou", 0.0),
                    "stability_rate":   coco.get("stability_rate", 0.0),
                }

                # Parse per-object data from SAM response
                for frame in sam_data.get("detections", []):
                    for det in frame.get("detections", []):
                        bbox = det.get("bbox", [0, 0, 0, 0])
                        run["objects"].append({
                            "object_id":      det.get("class_name", "object"),
                            "label":          "",           # filled by CLIP below
                            "bbox_x1":        bbox[0] if len(bbox) > 0 else "",
                            "bbox_y1":        bbox[1] if len(bbox) > 1 else "",
                            "bbox_x2":        bbox[2] if len(bbox) > 2 else "",
                            "bbox_y2":        bbox[3] if len(bbox) > 3 else "",
                            "sam_confidence": det.get("confidence", ""),
                            "clip_confidence": "",
                            "distance_cm":    det.get("distance_cm", ""),
                            "iou_score":      det.get("iou_with_previous", ""),
                            "is_stable":      det.get("is_stable_detection", ""),
                            "has_grasp":      False,
                            "grasp":          {},
                        })
                self.get_logger().info(f"  SAM: {run['sam']['total_detections']} objects detected")
            except Exception as e:
                self.get_logger().error(f"  SAM parse error: {e}")

        # Wait 500 ms for CLIP to auto-process the published SAM topic
        time.sleep(0.5)

        # ---- 2. CLIP ----
        self.get_logger().info(f"RUN #{run_no} — Step 2: CLIP /vision/classify_bbox_filtered")
        t_clip = time.perf_counter()
        clip_ok, clip_msg = self._call(self._clip_client, "/vision/classify_bbox_filtered")
        run["clip"]["latency_s"] = round(time.perf_counter() - t_clip, 3)
        if clip_ok and clip_msg:
            try:
                clip_data = json.loads(clip_msg)
                run["clip"] = {
                    "success":          True,
                    "latency_s":        run["clip"]["latency_s"],
                    "total_sam_regions": clip_data.get("total_sam_regions", 0),
                    "filtered_regions":  clip_data.get("filtered_regions", 0),
                }
                # Merge CLIP labels into objects by region_id index
                clip_regions = {r["region_id"]: r for r in clip_data.get("regions", [])}
                for idx, obj in enumerate(run["objects"]):
                    region = clip_regions.get(idx)
                    if region:
                        obj["label"]           = region.get("label", "")
                        obj["clip_confidence"] = region.get("confidence", "")
                self.get_logger().info(f"  CLIP: {run['clip']['filtered_regions']} regions classified")
            except Exception as e:
                self.get_logger().error(f"  CLIP parse error: {e}")

        # ---- 3. Scene Understanding ----
        self.get_logger().info(f"RUN #{run_no} — Step 3: Scene /vision/understand_scene")
        t_scene = time.perf_counter()
        scene_ok, scene_msg = self._call(self._scene_client, "/vision/understand_scene", timeout=15.0)
        run["scene"]["latency_s"] = round(time.perf_counter() - t_scene, 3)
        if scene_ok and scene_msg:
            try:
                scene_data = json.loads(scene_msg)
                run["scene"] = {
                    "success":           True,
                    "latency_s":         run["scene"]["latency_s"],
                    "scene_id":          scene_data.get("scene_id", ""),
                    "total_objects":     scene_data.get("total_objects", 0),
                    "total_relations":   scene_data.get("total_relations", 0),
                    "graspable_objects": scene_data.get("graspable_objects", 0),
                    "scene_description": scene_data.get("scene_description", ""),
                }

                # Per-object details (enriches existing objects list)
                scene_objects = scene_data.get("objects", {})
                for obj_id, obj_info in scene_objects.items():
                    # Try to match to an existing object entry by bounding box proximity
                    # or append a new entry if scene understanding found it
                    matched = False
                    s_bbox = obj_info.get("bbox", [])
                    for obj in run["objects"]:
                        if (len(s_bbox) >= 4 and
                                obj.get("bbox_x1") == s_bbox[0] and
                                obj.get("bbox_y1") == s_bbox[1]):
                            # Update with scene info
                            obj["label"]    = obj.get("label") or obj_info.get("label", "")
                            obj["has_grasp"] = obj_info.get("has_grasp", False)
                            if obj_info.get("has_grasp") and obj_info.get("grasp_quality") is not None:
                                obj["grasp"] = {
                                    "quality_score": obj_info.get("grasp_quality"),
                                    "width_m":       "",  # not available at this level
                                }
                            matched = True
                            break

                    if not matched:
                        s_bbox_safe = s_bbox if len(s_bbox) >= 4 else ["", "", "", ""]
                        run["objects"].append({
                            "object_id":      obj_id,
                            "label":          obj_info.get("label", ""),
                            "bbox_x1":        s_bbox_safe[0],
                            "bbox_y1":        s_bbox_safe[1],
                            "bbox_x2":        s_bbox_safe[2],
                            "bbox_y2":        s_bbox_safe[3],
                            "sam_confidence": "",
                            "clip_confidence": obj_info.get("confidence", ""),
                            "distance_cm":    obj_info.get("distance_cm", ""),
                            "iou_score":      "",
                            "is_stable":      "",
                            "has_grasp":      obj_info.get("has_grasp", False),
                            "grasp": {
                                "quality_score": obj_info.get("grasp_quality", ""),
                                "width_m":       "",
                            },
                        })

                    # Relations for this object
                    for rel in obj_info.get("relations", []):
                        run["relations"].append({
                            "subject":       obj_id,
                            "relation":      rel.get("relation", ""),
                            "target_object": rel.get("target_object", ""),
                            "confidence":    rel.get("confidence", ""),
                            "distance_2d":   rel.get("distance_2d", ""),
                            "description":   rel.get("description", ""),
                        })

                self.get_logger().info(
                    f"  Scene: {run['scene']['total_objects']} objects, "
                    f"{run['scene']['total_relations']} relations"
                )
            except Exception as e:
                self.get_logger().error(f"  Scene parse error: {e}")

        # ---- 4. OBB Angle Benchmark ----
        self.get_logger().info(f"RUN #{run_no} — Step 4: OBB /obb/find_object_angle")
        t_obb = time.perf_counter()
        obb_resp = self._call_obb()
        run["obb"]["latency_s"] = round(time.perf_counter() - t_obb, 3)

        if obb_resp is not None and obb_resp.success:
            run["obb"]["success"]       = True
            run["obb"]["total_objects"] = obb_resp.total_objects

            # Build a lookup: object_id -> OBB data
            obb_by_id = {}
            for i, oid in enumerate(obb_resp.object_ids):
                theta_rad  = obb_resp.thetas[i]
                # The service already stores the remapped angle (90 - geom_deg).
                # angle_deg is directly the display angle (0° = vertical).
                angle_deg  = math.degrees(theta_rad)
                obb_by_id[oid] = {
                    "obb_center_u":  obb_resp.centers_u[i],
                    "obb_center_v":  obb_resp.centers_v[i],
                    "obb_theta_rad": round(theta_rad, 5),
                    "obb_angle_deg": round(angle_deg, 2),
                    "obb_width_px":  round(obb_resp.widths[i], 2),
                    "obb_height_px": round(obb_resp.heights[i], 2),
                }

            # Merge into existing objects list (match by object_id)
            for obj in run["objects"]:
                obb = obb_by_id.get(obj.get("object_id", ""))
                if obb:
                    obj.update(obb)
                else:
                    # Initialise missing OBB fields so Excel has consistent columns
                    obj.setdefault("obb_center_u",  "")
                    obj.setdefault("obb_center_v",  "")
                    obj.setdefault("obb_theta_rad", "")
                    obj.setdefault("obb_angle_deg", "")
                    obj.setdefault("obb_width_px",  "")
                    obj.setdefault("obb_height_px", "")

            self.get_logger().info(
                f"  OBB: {obb_resp.total_objects} objects, "
                f"latency={run['obb']['latency_s']}s"
            )
        else:
            # Ensure OBB keys exist even if service was unavailable
            for obj in run["objects"]:
                for k in ("obb_center_u","obb_center_v","obb_theta_rad","obb_angle_deg","obb_width_px","obb_height_px"):
                    obj.setdefault(k, "")
            self.get_logger().warn("  OBB: service unavailable or returned failure")

        # ---- Finalize ----
        run["meta"]["latency_s"] = round(time.perf_counter() - t0, 3)
        self.get_logger().info(
            f"RUN #{run_no} complete in {run['meta']['latency_s']}s — "
            f"{len(run['objects'])} objects, {len(run['relations'])} relations"
        )
        self.get_logger().info("=" * 60)
        return run


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main(args=None):
    rclpy.init(args=args)

    node = VisionDataCollector()

    print()
    print("=" * 60)
    print("  Vision Pipeline Data Collector & Excel Exporter")
    print("=" * 60)
    print(f"  History file : {HISTORY_FILE}")
    print(f"  Excel file   : {EXCEL_FILE}")
    print(f"  Max runs kept: {MAX_HISTORY}")
    print("=" * 60)

    try:
        # Determine the next run number from existing history
        history = _load_history()
        last_run_no = history[-1]["meta"]["run_no"] if history else 0
        run_no = last_run_no + 1

        # Collect the run
        run = node.collect_run(run_no)

        # Append to history and persist
        history.append(run)
        _save_history(history)
        print(f"[OK] History saved → {HISTORY_FILE}  ({len(history[-MAX_HISTORY:])} runs stored)")

        # Export all stored runs to Excel
        _export_excel(history[-MAX_HISTORY:])

        print()
        print("Done. Open vision_runs_export.xlsx to view results.")
        print()

    except KeyboardInterrupt:
        pass
    finally:
        node.destroy_node()
        if rclpy.ok():
            rclpy.shutdown()


if __name__ == "__main__":
    main()
